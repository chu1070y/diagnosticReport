import os

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from module.common import Common

from openpyxl import Workbook


class Output(Common):
    def __init__(self):
        self.logger = self.get_logger()
        self.report_conf = self.get_config()['excel']

        # 그래프 저장 경로 설정
        self.save_dir = "./graphs"

    def fetch_data_from_mysql(self):
        self.logger.info("extract graph data from mysql")
        db_work = DBwork()
        table_name = self.get_config()['report']['graph_table_name']
        sql = f"select * from {db_work.db_name}.{table_name}"
        print('#### query:', sql)

        df = pd.read_sql(sql, db_work.db.get_connection())

        # 데이터가 tuple인 경우 DataFrame으로 변환
        if isinstance(df, list) or isinstance(df, tuple):
            # 열 이름 정의 (MySQL 컬럼 이름을 적절히 추가)
            columns = ["column1", "column2", "column3"]
            dataframe = pd.DataFrame(df, columns=columns)
        else:
            dataframe = df  # 이미 DataFrame인 경우

        return dataframe

    def create_excel(self, dataframe):
        self.logger.info("make excel data")
        output_file = self.report_conf['excel_file']

        # Workbook 생성
        wb = Workbook()

        # Sheet1에 데이터 저장
        ws_data = wb.active
        ws_data.title = "data"

        # 데이터 삽입
        for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx, column=c_idx, value=value)

        # 열 이름 삽입
        for col_idx, col_name in enumerate(dataframe.columns, start=1):
            ws_data.cell(row=1, column=col_idx, value=col_name)

        # Excel 파일 저장
        wb.save(output_file)

    def create_basicplot(self, dataframe):
        self.logger.info("make graph image files")

        os.makedirs(self.save_dir, exist_ok=True)

        dataframe["id"] = pd.to_datetime(dataframe["id"], format="%Y%m%d%H%M").astype(str)
        plt.set_loglevel('WARNING')

        for col in dataframe.columns:
            if col == "id":
                continue

            plt.figure(figsize=(8, 4))
            plot_line, = plt.plot(dataframe["id"], dataframe[col])

            num_ticks = 30  # 원하는 x축 표시 개수
            indices = np.linspace(0, len(dataframe) - 1, num_ticks, dtype=int)

            plt.title(col, color=plot_line.get_color(), fontweight="bold", fontsize=14, pad=20)
            plt.xticks(dataframe["id"].iloc[indices], rotation=90, fontsize=6)
            plt.grid(True, axis='y', linestyle='--', alpha=0.7)
            plt.ticklabel_format(axis='y', style='plain')
            plt.tight_layout()

            # 그래프 저장
            save_path = os.path.join(self.save_dir, f"{col}.png")
            plt.savefig(save_path, dpi=300)
            plt.close()

    def create_query_usage_chart(self, dataframe):
        self.logger.info("make query usage chart files")

        cols = ['Com_insert', 'Com_delete', 'Com_update', 'Com_select']
        vals = []

        for c in cols:
            vals.append(dataframe[c].sum())

        labels = [f"{col.split('_')[1]}\n({val / sum(vals) * 100:.1f}%)" for col, val in zip(cols, vals)]

        plt.figure(figsize=(8, 6))
        plt.pie(
            vals,
            explode=(0, 0, 0, 0),
            labels=labels,
            colors=['#ED7D31', '#FFC000', '#A5A5A5', '#4472C4'],
            # autopct='%1.1f%%',  # 퍼센트 표시
            wedgeprops={'edgecolor': 'white', 'linewidth': 2},
            labeldistance=1.1,
            # pctdistance=0.1,
            startangle=100,
            textprops={'fontsize': 12},
        )

        plt.title("Query Usages", fontsize=20, fontweight='bold', color="#4472C4", pad=20)

        plt.legend(
            cols,
            loc="lower center",
            bbox_to_anchor=(0.5, -0.05),
            ncol=4,
            fontsize=10,
            frameon=False
        )

        plt.tight_layout()

        save_path = os.path.join(self.save_dir, f"Query_usage.png")
        plt.savefig(save_path, dpi=300)
        plt.close()


if __name__ == "__main__":
    from function.db_work import DBwork

    e = Output()

    data = e.fetch_data_from_mysql()
    # e.create_excel(data)
    e.create_basicplot(data)
    e.create_query_usage_chart(data)

