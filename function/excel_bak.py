import pandas as pd

from module.common import Common

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter


class Excel(Common):
    def __init__(self):
        self.logger = self.get_logger()
        self.report_conf = self.get_config()['excel']

    def fetch_data_from_mysql(self):
        db_work = DBwork()
        table_name = self.get_config()['report']['graph_table_name']
        sql = f"select * from {db_work.db_name}.{table_name}"
        print('#### query:', sql)

        df = pd.read_sql(sql, db_work.db.get_connection())
        return df

    def create_excel_with_charts(self, data):
        output_file = self.report_conf['excel_file']

        # 데이터가 tuple인 경우 DataFrame으로 변환
        if isinstance(data, list) or isinstance(data, tuple):
            # 열 이름 정의 (MySQL 컬럼 이름을 적절히 추가)
            columns = ["column1", "column2", "column3"]
            dataframe = pd.DataFrame(data, columns=columns)
        else:
            dataframe = data  # 이미 DataFrame인 경우

        print(dataframe)

        # Workbook 생성
        wb = Workbook()

        # Sheet1에 데이터 저장
        ws_data = wb.active
        ws_data.title = "Sheet1"

        # 데이터 삽입
        for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws_data.cell(row=r_idx, column=c_idx, value=value)

        # 열 이름 삽입
        for col_idx, col_name in enumerate(dataframe.columns, start=1):
            ws_data.cell(row=1, column=col_idx, value=col_name)

        # 그래프 시트 생성
        ws_chart = wb.create_sheet(title="그래프")

        # 그래프 추가
        for col_idx in range(2, len(dataframe.columns) + 1):
            chart = BarChart()
            chart.title = dataframe.columns[col_idx - 1]
            chart.legend = None

            # 데이터 참조
            data = Reference(ws_data, min_col=col_idx, min_row=1, max_row=len(dataframe) + 1)
            categories = Reference(ws_data, min_col=1, min_row=2, max_row=len(dataframe) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)

            # X축 레이블 설정
            chart.x_axis.tickLblPos = 'low'  # 레이블을 아래쪽에 배치

            # 그래프를 시트에 배치
            col_letter = get_column_letter(col_idx)
            ws_chart.add_chart(chart, f"A{(col_idx - 1) * 15}")

        # Excel 파일 저장
        wb.save(output_file)


if __name__ == "__main__":
    from function.db_work import DBwork

    e = Excel()
    # db_work = DBwork()
    # e.create_excel_with_charts(db_work.get_graph_data())
    e.create_excel_with_charts(e.fetch_data_from_mysql())

